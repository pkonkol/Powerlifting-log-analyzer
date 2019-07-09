#import pandas as pd
from openpyxl import load_workbook
import re

XSL_FILE = 'data/program.xlsx'
UNITS = 'KG'

RPE_SCHEME = '((?:[1-9],[5])|(?:[1-9]|10))'
WEIGHT_SCHEME = '([0-9]+)'
REPS_SCHEME = '([0-9]+)'

class Training_spreadsheet_parser():
    pass


def parse_exercise(ws, col, row):
    e = Exercise()
    name = str(ws[col+row].value)
    log = str(ws[chr(ord(col)+1) + row].value)
    comment = str(ws[chr(ord(col)+1) + row].comment)

    print(name + " :: " + log)
    row = str(int(row)+1)
    return w,col, row

def parse_workout(ws, start_col, start_row, lrow, day):
    w = Workout(0, '', [], 0, )
    col, row = start_col, start_row
    date = str(ws[chr(ord(col)+1) + row].value)
    print("Day " + day + " at date " + date)

    while int(row) < int(lrow):
        col, row = parse_exercise(ws, col, row)

    col = chr(ord(col)+2)
    row = start_row
    return w, col, row

def parse_microcycle(ws, start_col, start_row):
    micro = Microcycle(0,0, [], "","")
    frow, lrow = 0, 0
    week = 0
    nweek = 0
    col, row = start_col, start_row
    m = None
    while not m:
        m = re.match("^W([0-9]+)$", str(ws[col+row].value))
        if m:
            frow = row
            week = m.group(1)
            week_comment = ws[col+row].comment
            print('week = ' + str(week))

            if int(week) == 0:
                return micro, start_col, lrow, week, nweek

            while True:
                row = str(int(row) + 1)
                m = re.match("^W([0-9]+)$", str(ws[col+row].value))
                if m:
                    nweek = m.group(1)
                    lrow = str(int(row) - 1)
                    break
            break
        else:
            row = str(int(row) + 1)

    col = chr(ord(col)+1)
    print(col)

    row=start_row
    m = re.match("^D([0-9]+)$", str(ws[col+row].value))
    while m:
        day = m.group(1)
        day_comment = ws[col+row].comment
        col, row = parse_workout(ws, col, row, lrow, day)
        m = re.match("^D([0-9]+)$", str(ws[col+row].value))
        print(m)

    print('micro ret: ' + lrow + ' ' + week)
    return micro, start_col, str(int(lrow)+1), week, nweek

def parse_mesocycle(ws, start_col, start_row):
    cur_col, cur_row = start_col, start_row
    meso = Mesocycle([], 0, 0, "")

    while True:
        microcycle, cur_col, cur_row, week, nweek = parse_microcycle(ws, cur_col, cur_row)
        print("|{},-{},; week-{},nweek-{}".format(cur_col, cur_row, week, nweek))

        meso.microcycles.append(microcycle)

        print(nweek)
        if nweek == 0:
            break

    return ":)"



class Exercise:
    def __init__(self, note=''):
        self.name = ''
        self.modifiers = []
        self.sets_planned = []
        self.sets_done = []
        self.note = note
        self.workset_volume = 0
        self.is_superset = False

    def workout_from_string(self, first_col_str, second_col_str): # First column contains exercise with modifiers and planned sets, second column contains done sets
        exercise_str, planned_str = first_col.split(':')
        self.name, self.modifiers = exercise_from_string(exercise_str)
        self.sets_planned = sets_planned_from_string(planned_str)
        self.sets_done = sets_done_from_string(second_col_str)

    def exercise_from_string(self, exercise_str):
        modifier_schemes = {' w/([a-zA-Z0-9]+)': 'with x',
                            ' t/(\d{4})': 'tempo XXXX',
                            ' wo/([a-zA-Z0-9]+)': 'without x',
                            ' p/([a-zA-z0-9]+)': 'movement patter mod'}

        modifiers = [(re.findall(key, name), value) for key, value in modifiers.items()]

        name = exercise_str
        for key in modifiers.keys():
            name = re.sub(key, '', name)

        return (name, modifiers)

    def sets_done_from_string(self, sets_str):
        set_scheme = {'reps': '', 'weight':'', 'RPE': '', 'set_no': ''}
        prior_weight = 0
        prior_reps = 0
        prior_rpe = 0


        schemes = {'^([0-9]+)x([0-9]+)@(([1-9],[5])|([1-9]|10))$': 1, #'weight x reps at RPE'
                   '^([0-9]+)@(([1-9],[5])|([1-9]|10))$': 2, #'weight at RPE (presumed same set number as planned)',
                   '^{weight}(?:@{rpe}){{2,}}$'.format(weight=WEIGHT_SCHEME, rpe= RPE_SCHEME): 3, #Multiple Weight@Rpe sets written in one string
                   '^{weight}x{reps}(?:@{rpe}){{2,}}$'.format(weight=WEIGHT_SCHEME,reps=REPS_SCHEME, rpe= RPE_SCHEME): 4, #Multiple WeightXReps@Rpe sets written in one string
                   }

        sets_str = re.split(' |;', sets_str)
        sets_done = []
        set_no = 1
        for set_str in sets_str:
            c_set = {'reps': '', 'weight':'', 'RPE': '', 'set_no': ''}
            while True:
                try:
                    match = [(re.match(key, set_str), value) for key, value in zip(schemes.keys(), schemes.values()) if re.match(key, set_str)][0]
                except IndexError:
                    print('Failed to match set with any of schemes')
                    print('Please enter correct string')
                    print(set_str, end='')
                    set_str = input()
                    continue
                break

            print('{}, {} - no{}, match={}, groups={}'.format(sets_str, set_str, set_no, match, match[0].groups()))
            if match[1] == 1:
                c_set['weight'] = float(match[0].group(1).replace(',', '.'))
                c_set['reps'] = int(match[0].group(2).replace(',', '.'))
                c_set['RPE'] = float(match[0].group(3).replace(',', '.'))
                c_set['set_no'] = set_no
                set_no += 1
                sets_done.append(c_set)
            elif match[1] == 2:
                c_set['weight'] = float(match[0].group(1).replace(',', '.'))
                c_set['RPE'] = float(match[0].group(2).replace(',', '.'))
                c_set['set_no'] = set_no
                set_no += 1
                sets_done.append(c_set)
            elif match[1] == 3:
                print('match[1]==3')
                multiset_weight = float(match[0].group(1).replace(',','.'))
                multiset_rpe_list = match[0].group(0).split('@')[1:]
                print('multiset_rpe_list: ' + str(multiset_rpe_list))
                for rpe in multiset_rpe_list:
                    c_set = {'reps': '', 'weight':'', 'RPE': '', 'set_no': ''}
                    c_set['weight'] = multiset_weight
                    c_set['RPE'] = float(rpe.replace(',','.'))
                    c_set['set_no'] = set_no
                    set_no += 1
                    sets_done.append(c_set)
            elif match[1] == 4:
                print('match[1]==4')
                multiset_weight = float(match[0].group(1).replace(',','.'))
                multiset_reps = int(match[0].group(2).replace(',', '.'))
                multiset_rpe_list = match[0].group(0).split('@')[1:]
                for rpe in multiset_rpe_list:
                    c_set = {'reps': '', 'weight':'', 'RPE': '', 'set_no': ''}
                    c_set['weight'] = multiset_weight
                    c_set['reps'] = multiset_reps
                    c_set['RPE'] = float(rpe.replace(',','.'))
                    c_set['set_no'] = set_no
                    set_no += 1
                    sets_done.append(c_set)

        return sets_done

    def sets_planned_from_string(self, sets_planned_str):
        sets_num = 0
        schemes = {'^x(\d+)@(([1-9],[5])|([1-9]|10))$': 'reps at RPE',
                   '^^([0-9]+)x([0-9]+)@([1-9][0-9]?|100)%$$': 'sets of reps at percentage',
                   '^([1-9][0-9]?|100)%@(([1-9],[5])|([1-9]|10))$': 'percentage at RPE',
                   '^x([0-9]+)(@(([1-9],[5])|([1-9]|10)))(@(([1-9],[5])|([1-9]|10)))+$': 'reps at RPE multiple', #needs furhter processing
                   '^([0-9]+)x([0-9]+)@(([1-9],[5])|([1-9]|10))$': 'sets of reps starting at RPE',
                   '^([0-9]+)x@(([1-9],[5])|([1-9]|10))$': 'number of sets at RPE',
                   '^([1-9][0-9]?|100)%x([0-9]+)$': 'reps at %1RM',
                   '^([0-9]+)x([0-9]+)@([1-9][0-9]?|100)%$': 'sets of reps at percentage',
                   '^ $': 'Sets x kilograms at RPE',
                   }

class Workout:
    def __init__(self, day, date, exercises, length, notes):
        self.day = day
        self.date = date
        self.exercises = exercises
        self.length = length
        self.note = note

class Microcycle:
    def __init__(self, date_start, date_end, workouts, drugs, notes):
        self.length = date_end - date_start
        self.date_s = date_start
        self.date_e = date_end
        self.workouts = workouts
        self.drugs = drugs
        self.notes = notes

class Mesocycle:
    def __init__(self, microcycles, date_start, date_end, notes):
        self.microcycles = microcycles
        self.date_start = date_start
        self.date_end = date_end
        self.notes = notes

class Bodyweight:
    pass


if __name__ == '__main__':
    wb = load_workbook(filename = XSL_FILE)
    ws = wb['program']

    parse_mesocycle(ws, 'A', '1')

