from collections import namedtuple
from enum import Enum
from utils import SheetCell, calculate_e1RM
from openpyxl import load_workbook
import copy
import datetime
import json
import logging
import re

MATCH_PLANNED_TO_DONE = False

RPE_SCHEME = '(?P<rpe>(?:[1-9](?:,|\.)[5])|(?:[1-9]|10)|(?:9\.\(3\)|9\.3|9\.\(6\)|9\.6))'
RPE_MULTISET_SCHEME = f'(?P<multi_rpe>(?:@{RPE_SCHEME}){{2,}})'
WEIGHT_SCHEME = '(?:(?P<weight>[0-9]+(?:\.[0-9]{1,3})?)' \
                '(?i)(?P<unit>kg|lbs|bw)?|(?P<bw>BW))'
PERCENTAGE_SCHEME = '(?P<percentage>[1-9][0-9]?|100)'
REPS_SCHEME = '(?P<reps>[0-9]+)'
SETS_SCHEME = '(?P<sets>[0-9]+)'

class WeightUnit(Enum):
    KG = 1
    LBS = 2
    BW = 3
    PERCENT_1RM = 4
    PERCENT_TOPSET = 5

logging.basicConfig(filename='pla.log',
    format='%(levelname)s:%(name)s:%(lineno)s  %(message)s',
    level=logging.DEBUG)
logger = logging.getLogger(__name__)

class TrainingSpreadsheetParser:
    """ Loads training spreadsheet from file and parses it into structure of
    training related objects, in a form easy for further processing
    """

    def __init__(self, file_dir, unit):
        self.unit = unit
        self.wb = load_workbook(filename = file_dir)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.mesocycles = []

    def get_mesocycles(self):
        return self.mesocycles

    def parse_worksheet(self):
        ws = self.ws
        current_cell = SheetCell(1,1)

        cval = ws.cell(row=current_cell.row, column=current_cell.col).value
        while cval != 'Sheet_end:':
            logger.debug(str(cval) + ' Mesocycle')
            if isinstance(cval, str) and re.match("^Mesocycle:", cval):
                m = self.parse_mesocycle(current_cell)
                self.mesocycles.append(m)

            current_cell.next_row()
            cval = ws.cell(row=current_cell.row, column=current_cell.col).value

    def parse_mesocycle(self, start_cell):
        ws = self.ws
        current_cell = copy.copy(start_cell)
        notes = ws.cell(row=current_cell.row, column=current_cell.col).comment
        date_range = self.ws.cell(row=current_cell.row,
                        column=current_cell.col+1).value
        name = re.sub('^Mesocycle:', '', self.ws.cell(
                      row=current_cell.row, column=current_cell.col).value
                     )
        logger.info(f'Parsing mesocycle: {name}')

        microcycles = []

        cval = ws.cell(row=current_cell.row, column=current_cell.col).value
        while cval != 'Mesocycle_end:':
            if isinstance(cval, str) and re.match("^W(?:\d+)", cval):
                logger.debug(cval + ' :Microcycle')
                micro = self.parse_microcycle(current_cell)
                logger.debug(f'{current_cell.col}-{current_cell.row};' \
                             f'week-{date_range}')
                microcycles.append(micro)

            current_cell.next_row()
            cval = ws.cell(row=current_cell.row, column=current_cell.col).value

        m = Mesocycle(microcycles, name, date_range, notes)
        return m

    def parse_microcycle(self, start_cell):
        ws = self.ws
        current_cell = copy.copy(start_cell)
        w_n = ws.cell(row=current_cell.row, column=current_cell.col).value
        notes = ws.cell(row=current_cell.row, column=current_cell.col).comment
        date = ws.cell(row=current_cell.row+1, column=current_cell.col).value
        workouts = []

        current_cell.next_col()
        cval = ws.cell(row=current_cell.row, column=current_cell.col).value
        while cval:
            if isinstance(cval, str) and re.match("^D(?:\d+)", cval):
                logger.debug(cval)
                workout = self.parse_workout(current_cell)
                workouts.append(workout)

            current_cell.next_col(); current_cell.next_col()
            cval = ws.cell(row=current_cell.row, column=current_cell.col).value

        m = Microcycle(date, workouts, '', notes)
        return m

    def parse_workout(self, current_cell):
        ws = self.ws
        current_cell = copy.copy(current_cell)
        exercises = []

        day = ws.cell(row=current_cell.row, column=current_cell.col).value
        date_place = ws.cell(row=current_cell.row,
                        column=current_cell.col+1).value
        notes = ws.cell(row=current_cell.row,
                    column=current_cell.col).comment

        current_cell.next_row()
        cval = ws.cell(row=current_cell.row, column=current_cell.col).value
        while cval != None:
            logger.debug(cval)
            planned_str = ws.cell(row=current_cell.row,
                              column=current_cell.col).value
            done_str = ws.cell(row=current_cell.row,
                           column=current_cell.col+1).value
            notes = ws.cell(row=current_cell.row,
                        column=current_cell.col).comment
            e = Exercise(planned_str, done_str, notes)
            exercises.append(e)

            current_cell.next_row()
            cval = ws.cell(row=current_cell.row, column=current_cell.col).value

        w = Workout(day, date_place, exercises, notes)
        return w


class Exercise:
    class SetType(Enum):
        NONE = -1
        RPE = 0
        WEIGHT = 1
        PERCENT_1RM = 2
        LOAD_DROP = 3
        FATIGUE_PERCENT = 4
        RPE_RAMP = 5

    DefaultUnit = WeightUnit.KG
    Weight = namedtuple('Weight', ('value', 'unit',))
    Set =  namedtuple('Set', ('type', 'reps', 'weight', 'rpe'))
    schemes_planned = (
        (re.compile(f'^x{REPS_SCHEME}@{RPE_SCHEME}$'), SetType.RPE), # x5@9 REPS at RPE
        (re.compile(f'^{SETS_SCHEME}x{REPS_SCHEME}@{PERCENTAGE_SCHEME}%$'),
            SetType.PERCENT_1RM),# 5x5@90% SETS of REPS at PERCENTAGE
        (re.compile(f'^{PERCENTAGE_SCHEME}%@{RPE_SCHEME}$'), # 80%@8 PERCENTAGE at RPE
            SetType.PERCENT_1RM),
        (re.compile(f'^x{REPS_SCHEME}{RPE_MULTISET_SCHEME}$'), # x5@8@9 REPS at RPE multiple #needs furhter processing
            SetType.RPE),
        (re.compile(f'^{SETS_SCHEME}x{REPS_SCHEME}\^@{RPE_SCHEME}$'),# 4x4^@7 SETS of REPS starting at RPE
            SetType.RPE),
        (re.compile(f'^{SETS_SCHEME}x@{RPE_SCHEME}$'), #  3x@9 number of SETS at RPE
            SetType.RPE),
        (re.compile(f'^{PERCENTAGE_SCHEME}%x{REPS_SCHEME}$'), # 80%x5 REPS at %1RM
            SetType.PERCENT_1RM),
        (re.compile(f'^{SETS_SCHEME}x{REPS_SCHEME}V{PERCENTAGE_SCHEME}%$'), # 3x3V90% SETS of REPS at PERCENTAGE
            SetType.LOAD_DROP),
        (re.compile(f'^x{REPS_SCHEME}\$@{RPE_SCHEME}$'), # 3x3$@9 SETS x WEIGHT at RPE
            SetType.RPE_RAMP),
        (re.compile(f'^{SETS_SCHEME}x{REPS_SCHEME}$'),# 3x8 SETS of REPS starting at RPE
            SetType.NONE),
        (re.compile(f'^x{REPS_SCHEME}@{RPE_SCHEME}\-{PERCENTAGE_SCHEME}%$'),# x4@9-7% SETS of REPS starting at RPE
            SetType.FATIGUE_PERCENT),
        (re.compile(f'^{WEIGHT_SCHEME}@{RPE_SCHEME}$'),# 160lbs@9 SETS of REPS starting at RPE
            SetType.RPE),
        (re.compile(f'^{WEIGHT_SCHEME}/{REPS_SCHEME}$'),# 150x5 SETS of REPS starting at RPE
            SetType.WEIGHT),
    )
    schemes_done = (
        (re.compile(f'^{WEIGHT_SCHEME}x{REPS_SCHEME}@{RPE_SCHEME}$'),#weightXreps at RPE
            SetType.WEIGHT),
        (re.compile(f'^{WEIGHT_SCHEME}@{RPE_SCHEME}$'), #'weight at RPE (presumed same set number as planned)',
            SetType.WEIGHT),
        (re.compile(f'^{WEIGHT_SCHEME}{RPE_MULTISET_SCHEME}$'), #Multiple Weight@Rpe sets written in one string
            SetType.WEIGHT),
        (re.compile(f'^{WEIGHT_SCHEME}x{REPS_SCHEME}' \
                    f'{RPE_MULTISET_SCHEME}$'), #Multiple WeightXReps@Rpe sets written in one string
            SetType.WEIGHT),
        (re.compile(f'^{SETS_SCHEME}x{REPS_SCHEME}' \
                    f'(?:\/|x|@){WEIGHT_SCHEME}$'),
            SetType.WEIGHT),
        (re.compile('^ *(?P<undone>X) *$'), # exercise not done
            SetType.NONE),
        (re.compile('^ *(?P<done>V) *$'),
            SetType.NONE),
        (re.compile(f'^(?P<multi_reps>(?:{REPS_SCHEME}(?:,|@))+){WEIGHT_SCHEME}$'), #Multiple sets at given weight
            SetType.WEIGHT),
        (re.compile(f'^{REPS_SCHEME}x{WEIGHT_SCHEME}$'), #Reps at weight
            SetType.WEIGHT),
    )


    def __init__(self, planned_str, done_str, notes):
        logger.debug(self.SetType.RPE)
        logger.debug(self.SetType)

        self.done = True
        self.is_superset = False
        self.workout_from_string(planned_str, done_str)
        self.notes = notes
        self.workset_volume = 0
        self.categories = {}

        if self.done:
            self.e1RM = self.get_e1RM()
        else:
            self.e1RM = 0

    def get_e1RM(self):
        if self.is_superset:
            x = [max([calculate_e1RM(ws['weight'], ws['reps'], ws['RPE'])
                for ws in e]) for e in self.sets_done]
            return x
        x = max([calculate_e1RM(ws['weight'], ws['reps'], ws['RPE'])
                for ws in self.sets_done])
        return x

    def workout_from_string(self, first_col_str, second_col_str):
    # First column contains exercise with modifiers and planned sets,
    # second column contains done sets
        logging.debug('first_col_str: ' + first_col_str)
        exercise_str, planned_str = first_col_str.split(':')
        if '&' in exercise_str:
            self.is_superset = True
            exercise_strs = exercise_str.split('&')
            planned_strs = planned_str.split('&')
            second_col_strs = second_col_str.split('&')
            second_col_strs = [s.strip() for s in second_col_strs]

            self.name, self.modifiers, self.sets_planned, self.sets_done = [],[],[],[]
            for e_str, p_str, s_c_str in zip(exercise_strs,
                                             planned_strs, second_col_strs):
                name, modifiers = self.exercise_from_string(e_str)
                self.name.append(e_str)
                self.modifiers.append(modifiers)
                self.sets_planned.append(self.sets_planned_from_string(p_str))
                self.sets_done.append(self.sets_done_from_string(s_c_str))
            return

        self.name, self.modifiers = self.exercise_from_string(exercise_str)
        self.sets_planned = self.sets_planned_from_string(planned_str)
        self.sets_done = self.sets_done_from_string(second_col_str)

    def exercise_from_string(self, exercise_str):
        modifier_schemes = (re.compile(' w/(?P<with>[a-zA-Z0-9]+)'), #'with x',
                            re.compile(' t/(?P<tempo>\d{4})'), #'tempo XXXX',
                            re.compile(' wo/(?P<without>[a-zA-Z0-9]+)'), #'without x',
                            re.compile(' p/(?P<pattern>[a-zA-z0-9]+)')) #'movement patter mod'

        name = exercise_str
        modifiers = [pattern.findall(name) for pattern in modifier_schemes]
        logging.debug(f'Modifiers: {modifiers}')

        for m in modifiers:
            for k in m:
                name = re.sub(f' \w+/{k}(?: |$)', ' ', name)

        return (name.rstrip(), modifiers)

    def sets_done_from_string(self, sets_str):
        schemes = self.schemes_done
        sets_str = re.split(' |;', sets_str)
        for set_str in sets_str:
            while True:
                try:
                    match = [(pattern[0].match(set_str), index+1, pattern[1])
                        for index, pattern
                        in enumerate(schemes)
                        if pattern[0].match(set_str)][0]
                except IndexError as e:
                    logging.exception(e)
                    print('Failed to match set with any of schemes')
                break
            logging.debug(f'Done: {sets_str}, {set_str}, ' \
                f'match={match}, groups={match[0].groups()}')

        sets_done = self.parse_matched_set(match)
        if sets_done == (0,):
            self.done = False
            sets_done = []
        elif sets_done == (1,):
            self.done = True
            sets_done = []

        return sets_done

    def sets_planned_from_string(self, sets_planned_str):
        schemes = self.schemes_planned

        sets_planned_str = sets_planned_str.strip()
        if sets_planned_str == '':
            return []
        logging.debug('sets planned str:' + sets_planned_str +'/')
        sets_str = re.split(' |;', sets_planned_str)
        logging.debug(sets_str)
        for set_str in sets_str:
            while True:
                try:
                    match = [(pattern[0].match(set_str), index+1, pattern[1])
                        for index, pattern
                        in enumerate(schemes)
                        if pattern[0].match(set_str)][0]
                except IndexError as e:
                    logging.exception(e)
                    print('Failed to match set with any of schemes')
                break

            logger.info(f'Planned: {sets_str}, {set_str}, ' \
                        f'match={match}, groups={match[0].groups()}')
        sets_planned = self.parse_matched_set(match)
        return sets_planned

    def parse_matched_set(self, match):
        groupdict = match[0].groupdict()
        if 'undone' in groupdict:
            return (0,)
        if 'done' in groupdict:
            return (1,)

        logger.debug(groupdict)
        sets = []

        getdict = {
            'set_type': match[2],
            'reps': int(match[0].group('reps').replace(',', '.')) if
                        'reps' in groupdict else None,
            'sets': int(match[0].group('sets').replace(',', '.')) if
                        'sets' in groupdict else None,
            'rpe': (float(match[0].group('rpe')
                    .replace(',', '.').replace('(','').replace(')',''))
                    if 'rpe' in groupdict else None),
            'weight': (float(match[0].group('weight').replace(',', '.'))
                    if 'weight' in groupdict and groupdict['weight'] != None else None),
            'percentage': (float(match[0].group('percentage').replace(',', '.'))/100.0
                            if 'percentage' in groupdict else None),
            'multi_rpe': match[0].group('multi_rpe').split('@')[1:] if
                         'multi_rpe' in groupdict else None,
            'multi_reps': match[0].group('multi_reps').strip('@').split(',') if
                          'multi_reps' in groupdict else None
        }
        logger.debug(getdict)

        if getdict['weight']:
            matched_unit = match[0].group('unit')
            if matched_unit == 'kg':
                getdict['unit'] = WeightUnit.KG
            elif matched_unit == 'lbs':
                getdict['unit'] = WeightUnit.LBS
            else:
                getdict['unit'] = self.DefaultUnit
        elif 'bw' in groupdict and groupdict['bw'] == 'BW':
            getdict['unit'] = WeightUnit.BW
        elif getdict['percentage']:
            if getdict['set_type']==self.SetType.PERCENT_1RM:
                getdict['unit'] = WeightUnit.PERCENT_1RM
            elif getdict['set_type']==self.SetType.LOAD_DROP:
                getdict['unit'] = -1
            elif getdict['set_type']==self.SetType.FATIGUE_PERCENT:
                getdict['unit'] = None
            else:
                raise StandardError('Unsupported Set Type for percentage')

        if 'sets' in groupdict:
            for i in range(0, getdict['sets']):
                _set = self.create_set_object(**getdict)
                sets.append(_set)

                if getdict['set_type'] == self.SetType.RPE:
                    if 'unit' in getdict and getdict['unit'] == WeightUnit.BW:
                        continue
                    getdict.update({'percentage': 1.0, 'unit': -1, 'rpe': None,
                                    'set_type': self.SetType.LOAD_DROP})
                    continue

                if getdict['set_type'] == self.SetType.LOAD_DROP:
                    getdict['unit'] -= 1
        elif 'multi_rpe' in groupdict:
            for rpe in getdict['multi_rpe']:
                getdict['rpe']=float(rpe.replace(',', '.')
                                         .replace('(','').replace(')',''))
                _set = self.create_set_object(**getdict)
                sets.append(_set)
        elif 'multi_reps' in groupdict:
            for reps in getdict['multi_reps']:
                getdict['reps'] = int(reps)
                _set = self.create_set_object(**getdict)
                sets.append(_set)
        else:
            _set = self.create_set_object(**getdict)
            sets.append(_set)

        return sets


    def create_set_object(self, set_type=None, reps=None, rpe=None,
                          weight=None, percentage=None, unit=None, **kwargs):
        w = self.Weight(weight if weight else
                        percentage if percentage else
                        None,
                unit)
        return self.Set(set_type, reps, w, rpe)

    def match_sets_planned_done(self):
        pass

    def volume(self):
        vol = 0.0
        if not self.is_superset:
            for s in self.sets_done:
                if s['reps'] and s['weight']:
                    vol += s['reps']*s['weight']
            return vol

    def str(self):
        if self.is_superset:
            sets_planned = [f"{s['weight'] or ''}x{s['reps'] or ''}" \
                f"@{s['RPE'] or ''}" if s!='&' else f' {s} '
                for e in self.sets_planned for s in e + ['&']]
            if self.done:
                sets_done = [f"{s['weight'] or ''}x{s['reps'] or ''}" \
                    f"@{s['RPE'] or ''}" if s!='&' else f' {s} '
                    for e in self.sets_done for s in e + ['&']]
            else:
                sets_done = "X"
            ret = (f"{'&'.join(self.name)}: {';'.join(sets_planned)}"\
                    f"| {';'.join(sets_done)} e1RM:" \
                    f"{';'.join(['{:4.2f}'.format(f) for f in self.e1RM])}",
                    0)
            return ret
        if self.done:
            sets_done = [f"{s['weight'] or ''}x{s['reps'] or ''}" \
                f"@{s['RPE'] or ''}" for s in self.sets_done]
        else:
            sets_done = "X"
        sets_planned = [f"{s['weight'] or ''}x{s['reps'] or ''}" \
                        f"@{s['RPE'] or ''}" for s in self.sets_planned]
        e1RM = self.e1RM if self.e1RM else 0.0
        return (f'{self.name}: {";".join(sets_planned)} ' \
                f'| {";".join(sets_done)}',
                e1RM)

class Workout:
    def __init__(self, day, date_place_str, exercises, notes):
        self.day = day
        self.date, self.place = self.parse_date_place(date_place_str)
        self.exercises = exercises
        self.notes = notes

    def calculate_volume(self):
        vol = 0.0
        for e in self.exercises:
            vol += e.volume()
        return vol

    def parse_date_place(self, date_place_str):
        if '@' in date_place_str:
            date, place = date_place_str.split('@')
        else:
            date, place = date_place_str, ''
        return (date, place)

    def str(self):
        return [e.str() for e in self.exercises]

class Microcycle:
    def __init__(self, date, workouts, drugs, notes):
        self.length = len(workouts) # TODO
        self.date_start, self.date_end = self.parse_date(date)
        self.workouts = workouts
        self.drugs = drugs
        self.notes = notes

    def parse_date(self,date_str):
        #TODO
        date_start = date_str
        date_end = date_str
        return date_start, date_end

    def calculate_volume(self):
        pass

    def str(self):
        return f'\tDate: {self.date_start}-{self.date_end}, notes: {self.notes}'


class Mesocycle:
    def __init__(self, microcycles, name, date_range, notes):
        self.name = name
        self.microcycles = microcycles
        self.date_start = date_range
        self.date_end = date_range
        self.notes = notes

    def str(self):
        return f'Name: {self.name}, Date:{self.date_start}-{self.date_end}, ' \
               f'Length: {len(self.microcycles)}, Notes: {self.notes}'

class ExerciseCategory:
    def __init__(self):
        DATASETS = 'datasets.json'
        with open(DATASETS, 'r') as ds:
            self.data = json.load(ds)


class Bodyweight:
    pass

if __name__ == '__main__':
    #wb = load_workbook(filename = XSL_FILE)
    #ws = wb['program']

    tsp = TrainingSpreadsheetParser(XSL_FILE, 'kg')
    tsp.parse_worksheet()

