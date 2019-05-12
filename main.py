import pandas as pd
from openpyxl import load_workbook
import re

xsl_file = 'data/program.xlsx'


def parse_exercise(ws, col, row):
    e = Exercise()
    name = str(ws[col+row].value)
    log = str(ws[chr(ord(col)+1) + row].value)
    comment = str(ws[chr(ord(col)+1) + row].comment)

    print(name + " :: " + log)
    row = str(int(row)+1)
    return w,col, row

def parse_workout(ws, start_col, start_row, lrow, day):
    w = Workout()
    col, row = start_col, start_row
    date = str(ws[chr(ord(col)+1) + row].value)
    print("Day " + day + " at date " + date)

    while int(row) < int(lrow):
        col, row = parse_exercise(ws, col, row)

    col = chr(ord(col)+2)
    row = start_row
    return w, col, row

def parse_microcycle(ws, start_col, start_row):
    micro = Microcycle(0,0, [], "","")
    frow, lrow = 0, 0
    week = 0
    nweek = 0
    col, row = start_col, start_row
    m = None
    while not m:
        m = re.match("^W([0-9]+)$", str(ws[col+row].value))
        if m:
            frow = row
            week = m.group(1)
            week_comment = ws[col+row].comment
            print('week = ' + str(week))

            if int(week) == 0:
                return micro, start_col, lrow, week, nweek

            while True:
                row = str(int(row) + 1)
                m = re.match("^W([0-9]+)$", str(ws[col+row].value))
                if m:
                    nweek = m.group(1)
                    lrow = str(int(row) - 1)
                    break
            break
        else:
            row = str(int(row) + 1)

    col = chr(ord(col)+1)
    print(col)

    row=start_row
    m = re.match("^D([0-9]+)$", str(ws[col+row].value))
    while m:
        day = m.group(1)
        day_comment = ws[col+row].comment
        col, row = parse_workout(ws, col, row, lrow, day)
        m = re.match("^D([0-9]+)$", str(ws[col+row].value))
        print(m)

    print('micro ret: ' + lrow + ' ' + week)
    return micro, start_col, str(int(lrow)+1), week, nweek

def parse_mesocycle(ws, start_col, start_row):
    cur_col, cur_row = start_col, start_row
    meso = Mesocycle([], 0, 0, "")

    while True:
        micro, cur_col, cur_row, week, nweek = parse_microcycle(ws, cur_col, cur_row)

        meso.microcycles.append(micro)

        print(nweek)
        if nweek == 0:
            break

    return ":)"



class Exercise:
    def __init__(self, name, modifiers, sets_planned, sets_done, note):
        self.name = name
        self.modifiers = modifiers
        self.sets_planned = sets_planned
        self.sets_done = sets_done
        self.note = note

class Workout:
    def __init__(self, day, date, exercises, length, note):
        self.day = day
        self.date = date
        self.exercises = exercises
        self.length = length
        self.note = note

class Microcycle:
    def __init__(self, date_start, date_end, workouts, drugs, notes):
        self.length = date_end - date_start
        self.date_s = date_start
        self.date_e = date_end
        self.workouts = workouts
        self.drugs = drugs
        self.notes = notes

class Mesocycle:
    def __init__(self, microcycles, date_start, date_end, notes):
        self.microcycles = microcycles
        self.date_start = date_start
        self.date_end = date_end
        self.notes = notes

wb = load_workbook(filename = xsl_file)

ws = wb['program']


parse_mesocycle(ws, 'A', '1')

